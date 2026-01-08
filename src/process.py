# =============================================================================
# 
# PROCESS FROM MAIN PANEL
# 
# Make sure to set the custom date range and to save the exported csv with
# a start and end date (inclusive) like this: yyyy-mm-dd yyyy-mm-dd.csv
# 
# =============================================================================

from __future__ import annotations
import datetime as datetime
from pathlib import Path
from typing import TextIO
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill

PATH_BASE = Path(__file__).parent
PATH_CONFIG = PATH_BASE / 'config'
PATH_INPUT = PATH_BASE / 'input'
PATH_OUTPUT = PATH_BASE / 'output'
PATH_TEMPLATES = PATH_BASE / 'templates'

PATH_VARIABLES = PATH_CONFIG / 'variables.txt'
PATH_TEMPLATE_FINAL_REPORT = PATH_TEMPLATES / 'final_report.xlsx'

FMT_DT_INPUT = '%Y-%m-%d %H-%M'
FMT_DT_OUTPUT = '%Y-%m-%d %H-%M'
FMT_DATE_OUTPUT = '%Y-%m-%d (%a)'

PLACEHOLDER_SKIP = '-'
NUMBER_BONUS = '--'

WEIGHT_XP = 1
WEIGHT_CONSISTENCY = 1
MAX_BONUS = 120

# Classes

class Student:
    name: str
    practices: set[Practice]

    def __init__(self: Student, name: str) -> None:
        self.name = name
        self.practices = set()

    def practices_between(self: Student, start: datetime.datetime, end: datetime.datetime) -> set[Practice]:
        return set(filter(lambda p: p.is_between(start, end), self.practices))
    
    def xp_between(self: Student, start: datetime.datetime, end: datetime.datetime) -> int:
        return sum(p.xp for p in self.practices_between(start, end))

    def __hash__(self: Student) -> int:
        return hash(self.name)

    def __repr__(self: Student) -> str:
        return self.name

class Practice:
    student: Student
    desc: str
    xp: int
    date: datetime.datetime

    def __init__(self: Practice, student: Student, desc: str, xp: int, date: datetime.datetime) -> None:
        self.student = student
        self.desc = desc
        self.xp = xp
        self.date = date

    def is_between(self: Practice, start: datetime.datetime, end: datetime.datetime) -> bool:
        return start <= self.date <= end

    def __hash__(self: Practice) -> int:
        return hash((self.student, self.desc, self.xp, self.date.strftime(FMT_DT_OUTPUT)))
    
    def __repr__(self: Practice) -> str:
        return f'{self.student} : {self.xp} @ {self.date.strftime(FMT_DT_OUTPUT)}'

class DuolingoMarker:
    students: dict[str, Student]
    aliases: dict[str, Student]
    skips: set[str]
    goal: int
    bonus_weeks: set[datetime.date]
    dates: set[datetime.date]

    def __init__(self: DuolingoMarker) -> None:
        self.students = {}
        self.aliases = {}
        self.skips = set()
        self.goal: 0
        self.bonus_weeks = set()
        self.dates = set()

    def parse_variables(self: DuolingoMarker):
        with open(PATH_VARIABLES, 'r') as f:
            lines = list(filter(lambda L: L and not L.startswith(';'), map(str.strip, f.readlines())))

            for line in lines:
                k, v = map(str.strip, line.split('::'))
                
                if k == 'goal':
                    self.goal = int(v)

                elif k == 'bonus week end':
                    y, m, d = map(int, v.split('-'))
                    sunday = datetime.date(y, m, d)
                    self.bonus_weeks.add(sunday) 

                elif k == 'alias':

                    alias, real = map(lambda s: s.strip().lower(), v.split('=='))

                    if not real:
                        real = alias

                    elif real == PLACEHOLDER_SKIP:
                        self.skips.add(alias)
                        continue

                    s = self.students.setdefault(real, Student(real))
                    self.aliases[alias] = s

    def parse_input_files(self: DuolingoMarker) -> None:        
        paths = PATH_INPUT.glob('*.csv')
        for path in paths:
            if path.stem.startswith('_'):
                continue

            self.parse_input_file(path)

    def parse_input_file(self: DuolingoMarker, path: Path) -> None:

        ts_start, ts_end = path.stem.split()
        dt_start = datetime.datetime.strptime(f'{ts_start} 00-00', FMT_DT_INPUT)
        dt_end = datetime.datetime.strptime(f'{ts_end} 11-59', FMT_DT_INPUT)

        self.dates.add(datetime.date(dt_start.year, dt_start.month, dt_start.day))
        self.dates.add(datetime.date(dt_end.year, dt_end.month, dt_end.day))

        with open(path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            next(reader)

            for row in reader:
                # Name, Username, Email, ..., ..., ..., ..., ..., ..., Total XP, ...rest

                alias = row[1]
                xp = int(row[10])
                dt = dt_start

                if alias.lower() in self.skips:
                    continue
                
                student = self.aliases[alias.lower()]
                desc = f'Main panel week summary {ts_start} to {ts_end}'

                practice = Practice(student, desc, xp, dt)
                student.practices.add(practice)
 
    def show_weeks(self: DuolingoMarker) -> None:
        weeks = self.get_weeks()
        if not weeks:
            print('No data found')
            return
        
        numbers = self.get_week_numbers(weeks)

        for i in range(1, len(weeks) + 1):
            week = weeks[-i]
            number = numbers[-i]
            print()
            print(self.format_week(*week, number))

            if i < len(weeks):
                choice = input('\nEnter to show another week or Q to quit: ').strip().upper()

                if choice == 'Q':
                    break
        
        print('\nFinished')

    def get_weeks(self: DuolingoMarker) -> list[tuple[datetime.date]]:
        if not self.dates:
            return
        
        weeks = []

        start = None

        for date in sorted(self.dates):

            # Started week?
            if start is None:
                start = date
                end = None

            # Sunday?
            if date.strftime('%w') == '0':
                end = date
                weeks.append((start, end))
                start = None
        
        # Didn't end on a Sunday?
        if end is None:
            end = date
            weeks.append((start, end))

        return weeks
    
    def get_week_numbers(self: DuolingoMarker, weeks: list[tuple[datetime.date]]) -> list[str]:
        n = 1
        numbers = []

        boni = self.bonus_weeks.copy()

        for (start, end) in weeks:
            for bonus in boni:
                if start <= bonus <= end:
                    numbers.append(NUMBER_BONUS)
                    boni.remove(bonus)
                    break

            else:
                numbers.append(f'{n:>2}')
                n += 1
        
        return numbers

    def format_week(self: DuolingoMarker, start: datetime.date, end: datetime.date, label: str='') -> None:
        start_dt = date_to_dt(start)
        end_dt = date_to_dt(end)

        if label:
            label += ' '

        s = f'Week {label}: {start.strftime(FMT_DATE_OUTPUT)} to {end.strftime(FMT_DATE_OUTPUT)}'

        s += "\n\nName".ljust(22) + ' : ' + "XP   : Counted\n"
        s += "=" * 37

        for stu in sorted(self.students.values(), key=lambda s: s.name):
            xp = stu.xp_between(start_dt, end_dt)

            name = stu.name.title().ljust(20)
            full = str(xp).ljust(4)
            capt = str(min(self.goal, xp)).ljust(3)

            s += '\n' + ' : '.join([name, full, capt])
        
        return s
    
    def calculate_final_report(self: DuolingoMarker) -> dict[str]:
        report = {
            'totals': {
                'xp': 0,
                'weeks': 0,
            },
            'averages': {
                'total xp': 0,
                'weekly xp': 0,
                '100% weeks': 0,
                '50% weeks': 0,
                'xp mark': 0,
                'consistency mark': 0,
            },
            'students': {}
        }

        weeks = self.get_weeks()
        numbers = self.get_week_numbers(weeks)

        report['totals']['weeks'] = sum(n is not NUMBER_BONUS for n in numbers)
        report['totals']['xp'] = self.goal * report['totals']['weeks']

        # Convert week dates to dts for practice comparisons...
        # Can't be done earlier because bonus uses dates...
        # TODO reconcile
        weeks = list((date_to_dt(s), date_to_dt(e)) for (s, e) in weeks)

        for stu in self.students.values():
            xps = []
            for (start_d, end_d) in weeks:
                xps.append(stu.xp_between(start_d, end_d))
            
            report['students'][stu.name] = self.get_student_stats(report['totals']['weeks'], xps)

        n = len(self.students)
        for key in report['averages']:
            report['averages'][key] = round(sum(s[key] for s in report['students'].values()) / n)

        return report
    
    def get_student_stats(self: DuolingoMarker, wks: int, xps: list[int]) -> dict[str, int]:
        """
        wks is the number of non-bonus weeks, which may differ from the length of the xp list.
        """
        d = {}

        # Bases
        xp_goal = self.goal * wks
        xp = sum(xps)

        # Stats
        d['total xp'] = xp
        d['weekly xp'] = round(xp / wks)
        d['100% weeks'] = sum(x >= self.goal for x in xps)
        d['50% weeks'] = sum(x >= (self.goal / 2) for x in xps) - d['100% weeks']

        # Marks
        # =MIN($B$5 / 100, MIN(I2 / $F$2, 1) + ((MAX(I2, $F$2) - $F$2) / ($B$2 * 100)))
        d['xp mark'] = min(
            MAX_BONUS / 100,
            min(xp / xp_goal, 1) + (max(xp, xp_goal) - xp_goal) / (self.goal * 100)
        )
        d['xp mark'] = round(100 * d['xp mark'])

        # =MIN($B$5 / 100, ( (I4/$F4) + (I5/$F5) ) / 2)
        d['consistency mark'] = min(
            MAX_BONUS / 100,
            (d['100% weeks'] / wks) + ((d['50% weeks'] / wks) / 2)
        )
        d['consistency mark'] = round(100 * d['consistency mark'])

        # Comments
        d['xp comment'] = f"Out of a goal of {xp_goal:,} XP, you earned {xp:,}. The weekly goal was {self.goal} and you earned an average of {d['weekly xp']} per week."
        d['consistency comment'] = self.format_consistency_comment(wks, d['100% weeks'], d['50% weeks'])

        return d
    
    def format_consistency_comment(self: DuolingoMarker, n: int, full: int, half: int) -> str:
        s = ""
        s += f"We did {n} weeks of practice."

        also = ""
        if full == 0:
            s += " You did not earn the full XP goal in any week."
        elif full == 1:
            s += f" You earned the full XP goal in 1 week."            
        else:
            s += f" You earned the full XP goal in {full} weeks."
            also = "also "
        
        if half == 1:
            s += f" You {also}earned at least half the XP goal in 1 week."
        elif half > 1:
            s += f" You {also}earned at least half the XP goal in {half} weeks."

        if (full + half) > n:
            s += f" (The total is higher than {n} because you also practiced in weeks when it wasn't required.)"

        return s
    
    def save_final_report(self: DuolingoMarker) -> None:
        report = self.calculate_final_report()

        # Workbook setup
        wb = openpyxl.load_workbook(PATH_TEMPLATE_FINAL_REPORT)
        ws = wb.active
        fill_name = PatternFill('solid', fgColor='153d64')
        font_name = Font(color='ffffff')

        # Totals
        totals = report['totals']
        ws['B2'] = totals['xp']
        ws['D2'] = totals['weeks']

        # Averages
        avgs = report['averages']
        ws['B3'] = avgs['total xp']
        ws['C3'] = avgs['weekly xp']
        ws['D3'] = avgs['100% weeks']
        ws['E3'] = avgs['50% weeks']
        ws['F3'] = avgs['xp mark']
        ws['H3'] = avgs['consistency mark']

        # Students
        for (i, stu) in enumerate(sorted(self.students.values(), key=lambda s: s.name)):
            data = report['students'][stu.name]
            r = i + 4
            
            ws[f'A{r}'] = stu.name
            ws[f'A{r}'].fill = fill_name
            ws[f'A{r}'].font = font_name

            ws[f'B{r}'] = data['total xp']
            ws[f'C{r}'] = data['weekly xp']
            ws[f'D{r}'] = data['100% weeks']
            ws[f'E{r}'] = data['50% weeks']
            ws[f'F{r}'] = data['xp mark']
            ws[f'G{r}'] = data['xp comment']
            ws[f'H{r}'] = data['consistency mark']
            ws[f'I{r}'] = data['consistency comment']

        try:
            wb.save(PATH_OUTPUT / 'final_report.xlsx')
            print(f'Saved report to {PATH_OUTPUT / "final_report.xlsx"}')
        except Exception as e:
            print('Could not save report')
            print(e)

    def format_final_report(self: DuolingoMarker) -> str:
        s = ""

        report = self.calculate_final_report()
        name_fill = len(max(self.students, key=len))

        # Totals
        totals = report['totals']
        s += '\nTOTALS\n'
        s += f"\nXP goal:      {totals['xp']} ({round(totals['xp'] / totals['weeks'])} per week)"
        s += f"\n# weeks goal: {totals['weeks']}"

        # Averages
        averages = report['averages']
        s += '\n\nAVERAGES\n'

        s += f"\nTotal XP earned:  {averages['total xp']}"
        s += f"\nWeekly XP earned: {averages['weekly xp']}"
        s += f"\n100% weeks:       {averages['100% weeks']}"
        s += f"\n50% weeks:        {averages['50% weeks']}"
        s += f"\nXP mark:          {averages['xp mark']}%"
        s += f"\nConsistency mark: {averages['consistency mark']}%"

        # Students

        s += "\n\nSTUDENT XP MARKS"
  
        for stu in sorted(self.students.values(), key=lambda s: s.name):
            data = report["students"][stu.name]
            s += f'\n{stu.name.ljust(name_fill)}\t{data["xp mark"]:>3}'
        
        s += "\n\nSTUDENT XP COMMENTS"
  
        for stu in sorted(self.students.values(), key=lambda s: s.name):
            data = report["students"][stu.name]
            s += f'\n\n{stu.name}\n{data["xp comment"]}'

        s += "\n\nSTUDENT CONSISTENCY MARKS"
  
        for stu in sorted(self.students.values(), key=lambda s: s.name):
            data = report["students"][stu.name]
            s += f'\n{stu.name.ljust(name_fill)}\t{data["consistency mark"]:>3}'
        
        s += "\n\nSTUDENT CONSISTENCY COMMENTS"
  
        for stu in sorted(self.students.values(), key=lambda s: s.name):
            data = report["students"][stu.name]
            s += f'\n\n{stu.name}\n{data["consistency comment"]}'

        return s
    
# Helpers
    
def date_to_dt(date: datetime.date) -> datetime.datetime:
    return datetime.datetime(date.year, date.month, date.day, 0, 0)

def dt_to_date(dt: datetime.datetime) -> datetime.date:
    return datetime.date(dt.year, dt.month, dt.day)
    
def clean_lines(f: TextIO) -> str:
    return map(str.lower, filter(None, map(str.strip, f.readlines())))

# Operations

def pick_student(d: DuolingoMarker) -> Student:
    choices = sorted(d.students)
    choice_str = ''
    for (i, name) in enumerate(choices):
        choice_str += f'{i + 1:>2}: {name}\n'
        
    print(f'Students:\n\n{choice_str}')
    number = int(input('Selection (enter number): '))
    return d.students[choices[number - 1]]

def make_marker() -> DuolingoMarker:
    d = DuolingoMarker()
    d.parse_variables()
    d.parse_input_files()
    return d

# Programs

def do_final_report() -> None:
    # TODO To be honest I'd rather output this to a spreadsheet

    print('Final report')
    d = make_marker()
    d.save_final_report()
    # print(d.format_final_report())
    input('\nPress Enter to exit')

def do_weekly_class_report() -> None:
    print('Weekly class report')
    d = make_marker()
    d.show_weeks()
    input('\nPress Enter to exit')

def do_weekly_student_report() -> None:
    print('Weekly student report')
    d = make_marker()
    s = pick_student(d)
    
    print(s)

    weeks = d.get_weeks()
    numbers = d.get_week_numbers(weeks)

    for (week, number) in reversed(list(zip(weeks, numbers))):
        start, end = week
        header = f'{number} {start.strftime(FMT_DATE_OUTPUT)} to {end.strftime(FMT_DATE_OUTPUT)}'
        xp = s.xp_between(date_to_dt(start), date_to_dt(end))
        print(f'{header}: {xp}')

def run():
    choice = input("Enter for weekly report, 1 for student report, 2 for final report: ").strip()

    if choice == "":
        do_weekly_class_report()
    elif choice == "1":
        do_weekly_student_report()
    elif choice == "2":
        do_final_report()

# Go

if __name__ == '__main__':
    run()
